"""Excel batch parser for organization import.

Parses .xlsx files with header row + N organization rows.
Supports Russian header synonyms and normalizes them.
"""

import re
from dataclasses import dataclass, field

HEADER_SYNONYMS: dict[str, list[str]] = {
    "organization_type": [
        "тип", "тип организации", "тип org", "type",
    ],
    "legal_name": [
        "наименование", "полное наименование", "организация",
        "название", "full name", "name",
    ],
    "short_name": [
        "краткое наименование", "короткое название", "short name",
    ],
    "inn": [
        "инн", "inn",
    ],
    "kpp": [
        "кпп", "kpp",
    ],
    "ogrn": [
        "огрн", "ogrn",
    ],
    "ogrnip": [
        "огрнип", "ogrnip",
    ],
    "legal_address": [
        "юридический адрес", "адрес", "юр. адрес", "юрид. адрес",
        "legal address",
    ],
    "actual_address": [
        "фактический адрес", "факт. адрес", "actual address",
    ],
    "residence_address": [
        "место жительства", "адрес регистрации", "прописка",
        "residence address",
    ],
    "director_name": [
        "директор", "руководитель", "генеральный директор",
        "head", "director",
    ],
    "phone": [
        "телефон", "phone", "tel",
    ],
    "email": [
        "email", "e-mail", "почта", "эл. почта",
    ],
    "bank_details": [
        "банковские реквизиты", "реквизиты", "банк", "bank details",
    ],
    "parent_inn": [
        "инн головной организации", "инн головной", "parent inn",
    ],
    "parent_kpp": [
        "кпп головной организации", "кпп головной", "parent kpp",
    ],
}


def _normalize_header(raw: str) -> str | None:
    """Normalize a raw header cell to a known field key."""
    cleaned = raw.strip().lower()
    cleaned = re.sub(r"\s+", " ", cleaned)
    for field_key, synonyms in HEADER_SYNONYMS.items():
        for synonym in synonyms:
            if cleaned == synonym:
                return field_key
    return None


def _normalize_value(value: str | None) -> str | None:
    """Trim whitespace, convert empty to None."""
    if value is None:
        return None
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    text = str(value).strip()
    return text if text else None


@dataclass
class ExcelRow:
    row_number: int
    raw_data: dict[str, str | None]
    normalized_data: dict[str, str | None] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)


def parse_xlsx_batch(content: bytes) -> tuple[list[str], list[ExcelRow]]:
    """Parse an xlsx file and return (unmapped_headers, rows).

    Returns:
        unmapped_headers: headers that could not be mapped to known fields
        rows: list of ExcelRow with raw_data and normalized_data
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl is required for Excel import") from exc

    wb = load_workbook(content, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        wb.close()
        return [], []

    header_map: dict[int, str] = {}
    unmapped: list[str] = []
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        raw_header = str(cell).strip()
        if not raw_header:
            continue
        mapped = _normalize_header(raw_header)
        if mapped:
            header_map[idx] = mapped
        else:
            unmapped.append(raw_header)

    result: list[ExcelRow] = []
    for row_num, row_data in enumerate(rows_iter, start=2):
        raw: dict[str, str | None] = {}
        normalized: dict[str, str | None] = {}
        has_content = False

        for col_idx, field_key in header_map.items():
            if col_idx < len(row_data):
                raw_val = row_data[col_idx]
                if isinstance(raw_val, str):
                    converted = raw_val
                elif raw_val is not None:
                    converted = str(raw_val)
                else:
                    converted = None
                raw[field_key] = _normalize_value(converted)
                if raw[field_key] is not None:
                    has_content = True

        if not has_content:
            continue

        normalized = dict(raw)

        row = ExcelRow(row_number=row_num, raw_data=raw, normalized_data=normalized)
        result.append(row)

    wb.close()
    return unmapped, result
